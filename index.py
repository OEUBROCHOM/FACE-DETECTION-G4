import tkinter as tk
from tkinter import messagebox, filedialog
import face_recognition
import cv2
import numpy as np
import openpyxl
import os
import threading
import pyttsx3
import datetime

# File path for user data (Excel)
USER_DATA_FILE = "user_data.xlsx"

def load_user_data():
    user_data = {}
    if os.path.exists(USER_DATA_FILE):
        workbook = openpyxl.load_workbook(USER_DATA_FILE)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            if len(row) < 6:
                continue  # Skip rows with missing values
            name, encoding_str, password, student_class, student_id, attendance = row
            try:
                encoding = np.fromstring(encoding_str, sep=' ')  # Convert the string encoding back into a list
                user_data[name] = {
                    "name": name,
                    "encoding": encoding,
                    "password": password,
                    "class": student_class,
                    "id": student_id,
                    "attendance": attendance == 'True'  # Convert 'True'/'False' to boolean
                }
            except Exception as e:
                print(f"Error processing row for {name}: {e}")
    return user_data


# Save user data to Excel
def save_user_data():
    if not os.path.exists(USER_DATA_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["name", "encoding", "password", "class", "id", "attendance"])  # Add headers
    else:
        workbook = openpyxl.load_workbook(USER_DATA_FILE)
        sheet = workbook.active

    for name, data in user_data.items():
        encoding_str = ' '.join(map(str, data['encoding']))  # Convert the encoding array to a string
        sheet.append([data['name'], encoding_str, data['password'], data['class'], data['id'], str(data['attendance'])])
    
    workbook.save(USER_DATA_FILE)

# Global variable to store user data
user_data = load_user_data()

def store_user_in_memory(name, face_encoding, password, student_class, student_id):
    user_data[name] = {
        "name": name,
        "encoding": face_encoding,
        "password": password,
        "class": student_class,
        "id": student_id,
        "attendance": False  # Initially, attendance is not marked
    }
    save_user_data()
    messagebox.showinfo("Success", f"User {name} has been added successfully.")

def validate_password(name, password):
    user_info = user_data.get(name, {})
    return user_info.get("password") == password

def upload_photo():
    file_path = filedialog.askopenfilename(
        title="Select a photo",
        filetypes=[("Image Files", "*.jpg;*.jpeg;*.png")]
    )
    if not file_path:
        messagebox.showinfo("Info", "No photo selected.")
        return

    user_name = name_entry.get()
    password = password_entry.get()
    student_class = class_entry.get()
    student_id = id_entry.get()

    if not user_name or not password or not student_class or not student_id:
        messagebox.showerror("Error", "Please enter all fields before uploading a photo.")
        return

    try:
        image = face_recognition.load_image_file(file_path)
        face_encodings = face_recognition.face_encodings(image)

        if face_encodings:
            store_user_in_memory(user_name, face_encodings[0], password, student_class, student_id)
        else:
            messagebox.showerror("Error", "No face detected in the uploaded photo. Please try again with a clear photo.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while processing the photo: {str(e)}")

def start_scan():
    messagebox.showinfo("Scan", "Starting live face scan... Please look at the camera.")
    threading.Thread(target=live_face_recognition, daemon=True).start()
def live_face_recognition():
    video_capture = cv2.VideoCapture(0)

    # Set the camera resolution
    video_capture.set(cv2.CAP_PROP_FRAME_WIDTH, 1400)
    video_capture.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)


    # Create the OpenCV window
    window_name = "Live Face Recognition - Press 'q' to quit"
    cv2.namedWindow(window_name)  # Ensure the window exists

    known_encodings = [np.array(data["encoding"]) for data in user_data.values()]
    known_names = list(user_data.keys())

    frame_count = 0
    exit_button_pressed = False  # Track if the exit button was pressed
    welcomed_users = set()  # To store users who have been welcomed already
    
    while not exit_button_pressed:
        ret, frame = video_capture.read()
        if not ret:
            print("Error: Failed to access the webcam.")
            break

        frame_count += 1
        if frame_count % 5 != 0:
            continue

        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
            matches = face_recognition.compare_faces(known_encodings, face_encoding, tolerance=0.6)
            name = "Unknown"
            attendance = "Not marked"

            face_distances = face_recognition.face_distance(known_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)
            if matches[best_match_index]:
                name = known_names[best_match_index]
                if not user_data[name]["attendance"]:
                    user_data[name]["attendance"] = True
                    save_user_data()
                    attendance = "Marked"
                else:
                    attendance = "Already marked"

                # If the user is recognized and hasn't been welcomed already, speak their name
                if name not in welcomed_users:
                    speak_welcome_message(f"Welcome to {name}!")
                    welcomed_users.add(name)  # Add to welcomed users to prevent repeated messages

            top *= 4
            right *= 4
            bottom *= 4
            left *= 4

            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)

        # Draw an Exit button on the video feed
        button_top_left = (50, 50)
        button_bottom_right = (200, 100)
        cv2.rectangle(frame, button_top_left, button_bottom_right, (0, 0, 255), -1)  # Filled rectangle for the button
        cv2.putText(frame, "Exit", (70, 85), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)

        # Detect mouse click on the button
        def on_mouse_click(event, x, y, flags, param):
            nonlocal exit_button_pressed
            if event == cv2.EVENT_LBUTTONDOWN:
                if button_top_left[0] <= x <= button_bottom_right[0] and button_top_left[1] <= y <= button_bottom_right[1]:
                    exit_button_pressed = True

        cv2.setMouseCallback(window_name, on_mouse_click)

        cv2.imshow(window_name, frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):  # Exit if 'q' is pressed
            break

    video_capture.release()
    cv2.destroyAllWindows()
# .............................................................................................
# File path for attendance logs
ATTENDANCE_LOG_FILE = "attendance_logs.xlsx"

# Define the check-in cutoff and check-out cutoff times
CHECK_IN_CUTOFF = datetime.time(7, 30)  # 7:30 AM
CHECK_OUT_CUTOFF = datetime.time(17, 0)  # 5:00 PM

# Initialize attendance log file
def initialize_attendance_log():
    if not os.path.exists(ATTENDANCE_LOG_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Student ID", "Class", "Date", "Time", "Status", "Remarks"])  # Add headers
        workbook.save(ATTENDANCE_LOG_FILE)

# Log attendance to the Excel file with remarks for Late, On Time, and Overtime
def log_attendance(name, student_id, student_class, status, remarks=None):
    current_time = datetime.datetime.now()
    date_str = current_time.strftime("%Y-%m-%d")
    time_str = current_time.strftime("%H:%M:%S")

    if os.path.exists(ATTENDANCE_LOG_FILE):
        workbook = openpyxl.load_workbook(ATTENDANCE_LOG_FILE)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Append attendance with remarks (Late, On Time, or Overtime)
    sheet.append([name, student_id, student_class, date_str, time_str, status, remarks])
    workbook.save(ATTENDANCE_LOG_FILE)

# Check-In or Check-Out functionality
def process_attendance(check_in=True):
    action = "Check-In" if check_in else "Check-Out"
    messagebox.showinfo(action, f"Starting {action} face scan. Please look at the camera.")
    threading.Thread(target=live_face_recognition_attendance, args=(check_in,), daemon=True).start()

def live_face_recognition_attendance(check_in=True):
    video_capture = cv2.VideoCapture(0)
    known_encodings = [np.array(data["encoding"]) for data in user_data.values()]
    known_names = list(user_data.keys())

    action = "Check-In" if check_in else "Check-Out"
    recognized_user = None

    while True:
        ret, frame = video_capture.read()
        if not ret:
            print("Error: Unable to access the camera.")
            break

        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

        for face_encoding in face_encodings:
            matches = face_recognition.compare_faces(known_encodings, face_encoding)
            if True in matches:
                first_match_index = matches.index(True)
                recognized_user = known_names[first_match_index]
                user_info = user_data[recognized_user]

                current_time = datetime.datetime.now()
                remarks = None  # Default to no remarks

                if check_in:
                    if not user_info["attendance"]:
                        user_info["attendance"] = True
                        save_user_data()

                        # Check if it's a late check-in
                        if current_time.time() > CHECK_IN_CUTOFF:
                            remarks = "Late"
                        # Check if it's exactly on time
                        elif current_time.time() <= CHECK_IN_CUTOFF:
                            remarks = "On Time"

                        log_attendance(
                            name=user_info["name"],
                            student_id=user_info["id"],
                            student_class=user_info["class"],
                            status="Check-In",
                            remarks=remarks
                        )
                        speak_welcome_message(f"{recognized_user} Checked In successfully!")
                    else:
                        messagebox.showinfo("Info", f"{recognized_user} has already Checked-In.")
                else:
                    if user_info["attendance"]:
                        user_info["attendance"] = False
                        save_user_data()

                        # Check if it's an early check-out (before 5:00 PM)
                        if current_time.time() < CHECK_OUT_CUTOFF:
                            remarks = "Leave early"
                        # Check if it's exactly on time (after 5:00 PM)
                        elif current_time.time() >= CHECK_OUT_CUTOFF:
                            remarks = "On Time"

                        log_attendance(
                            name=user_info["name"],
                            student_id=user_info["id"],
                            student_class=user_info["class"],
                            status="Check-Out",
                            remarks=remarks
                        )
                        speak_welcome_message(f"{recognized_user} Checked Out successfully!")
                    else:
                        messagebox.showinfo("Info", f"{recognized_user} has not Checked-In yet.")

                video_capture.release()
                cv2.destroyAllWindows()
                return

        cv2.imshow("Face Recognition", frame)
        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

    video_capture.release()
    cv2.destroyAllWindows()

# Initialize attendance log file
initialize_attendance_log()

# .................................................................................................
# Initialize text-to-speech engine
tts_engine = pyttsx3.init()

def speak_welcome_message(message):
    tts_engine.say(message)
    tts_engine.runAndWait()

def submit_form():
    user_name = name_entry.get().strip()
    password = password_entry.get().strip()
    student_class = class_entry.get().strip()
    student_id = id_entry.get().strip()

    if not user_name or not password or not student_class or not student_id:
        messagebox.showerror("Error", "Please enter all fields.")
        return


    if user_name in user_data:
        if validate_password(user_name, password):
            welcome_message = f"Welcome back, {user_name}! Starting face scan now."
            messagebox.showinfo("Welcome", welcome_message)
            speak_welcome_message(welcome_message)  # Speak the welcome message
            # Start face scan
            threading.Thread(target=live_face_recognition, daemon=True).start()
        else:
            messagebox.showerror("Error", "Invalid password.")
            return
    else:
        register = messagebox.askyesno("Register", "User not found. Do you want to register?")
        if register:
            capture_face(user_name, password, student_class, student_id)
        return
# Back to Welcome Screen
def show_welcome_screen():
    login_frame.pack_forget()
    welcome_frame.pack(pady=100)

# Capture Face for Registration
def capture_face(user_name, password, student_class, student_id):
    video_capture = cv2.VideoCapture(0)
    messagebox.showinfo("Info", "Look at the camera to capture your face.")
    captured = False

    while not captured:
        ret, frame = video_capture.read()
        if not ret:
            messagebox.showerror("Error", "Failed to access the webcam.")
            break

        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_encodings = face_recognition.face_encodings(rgb_frame)

        if face_encodings:
            store_user_in_memory(user_name, face_encodings[0], password, student_class, student_id)
            captured = True
            messagebox.showinfo("Success", f"Face captured and user {user_name} registered successfully.")

        cv2.imshow("Capturing Face - Press 'q' to quit", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    video_capture.release()
    cv2.destroyAllWindows()

# GUI Setup
def show_login_screen():
    welcome_frame.pack_forget()
    login_frame.pack(pady=20)

def start_scan_directly():
    welcome_frame.pack_forget()
    start_scan()

# Help Button Functionality
def show_help():
    help_window = tk.Toplevel(root)
    help_window.title("Help Guide")
    help_window.geometry("500x300")
    help_window.configure(bg="#34495e")
    
    # Help instructions
    instructions =              """Welcome to the Face Recognition App!
    
1. To register, provide your name, password, class, and student ID.
2. Upload a photo of your face to be registered.
3. To login, enter your name and password.
4. After logging in, the app will start scanning your face.
5. Attendance is automatically marked upon successful recognition.
6. For any issues, please contact support.

                                Press 'OK' to close this window.
"""
    label = tk.Label(help_window, text=instructions, font=("Arial", 12), bg="#34495e", fg="white", justify="left", anchor="w")
    label.pack(padx=10, pady=20)
    
    # OK button to close the Help window
    ok_button = tk.Button(help_window, text="OK", command=help_window.destroy, font=("Arial", 14, "bold"), bg="blue", fg="white", bd=5, relief="raised",width=10)
    ok_button.pack(pady=10)

# Exit Button Functionality
def exit_application():
    root.destroy()  # Closes the main application window


root = tk.Tk()
root.title("Face Recognition App")
root.geometry("800x800")
root.configure(bg="#34495e")


# Welcome Screen
welcome_frame = tk.Frame(root, bg="#34495e")
welcome_label = tk.Label(welcome_frame, text="Welcome to Face Recognition App", font=("Arial", 24, "bold"), bg="#34495e", fg="white")
welcome_label.pack(pady=50)

login_button = tk.Button(welcome_frame, text="Login or Register", command=show_login_screen, font=("Arial", 14, "bold"), bg="blue", fg="white", bd=5, relief="raised", width=20)
login_button.pack(pady=20)

scan_button = tk.Button(welcome_frame, text="Scan", command=start_scan_directly, font=("Arial", 14, "bold"), bg="green", fg="white", bd=5, relief="raised", width=20)
scan_button.pack(pady=10)


# Frame for Check-In and Check-Out buttons
check_frame = tk.Frame(welcome_frame, bg="#34495e")  # Create a sub-frame for horizontal alignment
check_frame.pack(pady=10)  # Place it below the Scan button with some padding

check_in_button = tk.Button(check_frame, text="Check-In", font=("Arial", 14), bg="green", fg="white", command=lambda: process_attendance(check_in=True),  width=15)
check_in_button.pack(side="left", padx=10)  # Align left with padding

check_out_button = tk.Button(check_frame, text="Check-Out", font=("Arial", 14), bg="blue", fg="white", command=lambda: process_attendance(check_in=False), width=15)
check_out_button.pack(side="left", padx=10)  # Align left with padding


# Adding Exit Button to Welcome Screen
exit_button = tk.Button(welcome_frame, text="Exit", command=exit_application, font=("Arial", 14, "bold"), bg="red", fg="white", bd=5, relief="raised", width=20)
exit_button.pack(side="left",pady=10,padx=10) 

# Help Button
help_button = tk.Button(welcome_frame, text="Help!", command=show_help, font=("Arial", 14, "bold"), bg="white", fg="red", bd=5, relief="raised", width=20)
help_button.pack(side="left",pady=10,padx=10)

welcome_frame.pack(pady=100)

# Login/Register Screen
login_frame = tk.Frame(root, bg="#34495e")

header_label = tk.Label(login_frame, text="Login or Register", font=("Arial", 30, "bold"), bg="#34495e", fg="white")
header_label.pack(pady=20)

name_label = tk.Label(login_frame, text="Name:", font=("Arial", 14, "bold"), bg="#34495e", fg="white", width=20)
name_label.pack(pady=5)
name_entry = tk.Entry(login_frame, font=("Arial", 12), width=30, bd=2, relief="groove")
name_entry.pack(pady=10)

password_label = tk.Label(login_frame, text="Password:", font=("Arial", 14, "bold"), bg="#34495e", fg="white")
password_label.pack(pady=5)
password_entry = tk.Entry(login_frame, font=("Arial", 12), show="*", width=30, bd=2, relief="groove")
password_entry.pack(pady=10)

class_label = tk.Label(login_frame, text="Class:", font=("Arial", 14, "bold"), bg="#34495e", fg="white")
class_label.pack(pady=5)
class_entry = tk.Entry(login_frame, font=("Arial", 12), width=30, bd=2, relief="groove")
class_entry.pack(pady=10)

id_label = tk.Label(login_frame, text="Student ID:", font=("Arial", 14, "bold"), bg="#34495e", fg="white")
id_label.pack(pady=5)
id_entry = tk.Entry(login_frame, font=("Arial", 12), width=30, bd=2, relief="groove")
id_entry.pack(pady=10)

upload_button = tk.Button(login_frame, text="Upload Photo", command=upload_photo, font=("Arial", 14, "bold"), bg="blue", fg="white", bd=5, relief="raised", width=15)
upload_button.pack(pady=10)

# Submit and Back buttons in the same row
button_frame = tk.Frame(login_frame, bg="#34495e")  # A new frame to hold the buttons side by side
button_frame.pack(pady=20)

back_button = tk.Button(button_frame, text="Back", command=show_welcome_screen, font=("Arial", 14, "bold"), bg="red", fg="white", bd=5, relief="raised", width=20)
back_button.pack(side="left", padx=10)

submit_button = tk.Button(button_frame, text="Submit", command=submit_form, font=("Arial", 14, "bold"), bg="green", fg="white", bd=5, relief="raised", width=20)
submit_button.pack(side="left", padx=10)


login_frame.pack(pady=20)  # Pack the login_frame at the end

root.mainloop()
